#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MotorNoLoadTester - Industrial Motor Testing System

Version: 1.3.0
Author: Sanjev Kumar (SK96)
License: Apache License 2.0

Copyright 2026 Sanjev Kumar

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

import hashlib
import os
import re
import struct
import subprocess
import sys
import tempfile
import threading
from datetime import datetime
from time import sleep

import keyboard
import pandas as pd
import requests
from colorama import Fore, Style, init
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from prettytable import PrettyTable  # type: ignore
from pymodbus.client.sync import ModbusSerialClient

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk


APP_VERSION = "1.3.0"
GITHUB_API = "https://api.github.com/repos/sanjevkumar961/Barcode-Scanning-MFM485-Meter-Modbus-Integration/releases/latest"
APP_NAME = "MotorNoLoadTester.exe"
UPDATER_NAME = "updater.exe"


def is_online() -> bool:
    """Check if internet connection is available."""
    try:
        requests.get("https://api.github.com", timeout=3)
        return True
    except requests.RequestException:
        return False


def sha256(path: str) -> str:
    """Calculate SHA256 hash of a file."""
    progress = DownloadProgress("Verifying Update")
    progress.set_text("Verifying file integrity...")
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for b in iter(lambda: f.read(8192), b""):
            h.update(b)
    progress.close()
    return h.hexdigest()


def check_for_update() -> None:
    """Check for available updates from GitHub and download if approved."""
    if not is_online():
        return  # Offline → do nothing

    try:
        r = requests.get(GITHUB_API, timeout=5)
        r.raise_for_status()
        data = r.json()

        latest_version = data["tag_name"].lstrip("v")

        if latest_version <= APP_VERSION:
            return

        confirm = messagebox.askyesno(
            "Update Available",
            f"Version {latest_version} is available.\n\nDo you want to update now?",
        )

        if not confirm:
            return

        exe_url = None
        sha_url = None

        for asset in data["assets"]:
            if asset["name"].endswith(".exe"):
                exe_url = asset["browser_download_url"]
            elif asset["name"].endswith(".sha256"):
                sha_url = asset["browser_download_url"]

        if not exe_url or not sha_url:
            return

        temp_dir = tempfile.mkdtemp()
        new_exe = os.path.join(temp_dir, f"MotorNoLoadTester_update_{latest_version}.exe")
        sha_file = os.path.join(temp_dir, f"MotorNoLoadTester_update_{latest_version}.sha256")

        download_with_progress(exe_url, new_exe)

        with requests.get(sha_url) as r:
            with open(sha_file, "w") as f:
                f.write(r.text.strip())

        expected = open(sha_file).read().strip()
        if sha256(new_exe) != expected:
            messagebox.showerror("Update Failed", "File integrity check failed.")
            return

        subprocess.Popen([UPDATER_NAME, new_exe, sys.executable])
        sys.exit()

    except (requests.RequestException, KeyError, IOError):
        # Silent fail (enterprise-safe)
        return


class DownloadProgress:
    def __init__(self, title="Downloading Update"):
        self.root = tk.Tk()
        self.root.title(title)
        self.root.resizable(False, False)

        self.label = tk.Label(self.root, text="Preparing download...")
        self.label.pack(padx=20, pady=(15, 5))

        self.progress = ttk.Progressbar(
            self.root, orient="horizontal", length=300, mode="determinate"
        )
        self.progress.pack(padx=20, pady=10)

        self.percent = tk.Label(self.root, text="0%")
        self.percent.pack(pady=(0, 15))

        self.root.protocol("WM_DELETE_WINDOW", lambda: None)

    def update(self, downloaded, total):
        percent = int(downloaded * 100 / total)
        self.progress["value"] = percent
        self.percent.config(text=f"{percent}%")
        self.root.update_idletasks()

    def set_text(self, txt):
        self.label.config(text=txt)
        self.root.update_idletasks()

    def close(self):
        self.root.destroy()


def download_with_progress(url: str, output_path: str) -> None:
    """Download a file with progress indication."""
    progress_ui = DownloadProgress()

    with requests.get(url, stream=True) as r:
        r.raise_for_status()
        total = int(r.headers.get("content-length", 0))
        downloaded = 0

        progress_ui.set_text("Downloading update...")

        with open(output_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total > 0:
                        progress_ui.update(downloaded, total)

    progress_ui.set_text("Download complete")
    progress_ui.close()


def getSheetLastRow(serial_name: str) -> None:
    """Get the last used row in the sheet for a given serial name."""
    global excel_row, sheet

    # Explicitly target the serial_name sheet
    if serial_name in book.sheetnames:
        sheet = book[serial_name]
    else:
        # Handle the case where serial_name sheet does not exist
        sheet = book.create_sheet(serial_name)
        heading()

    # Find the last used row in column A (S.No column)
    last_row = 1
    while sheet.cell(row=last_row, column=1).value is not None:
        last_row += 1

    # Set excel_row to the next available row
    excel_row = last_row - 1


def getConfig(key_name: str) -> dict | None:
    """Get configuration value from Config sheet."""
    global config_sheet
    # Check if the "Config" sheet exists
    if "Config" in book.sheetnames:
        # Access the "Config" sheet
        config_sheet = book["Config"]
    else:
        # Create the "Config" sheet and add headers
        config_sheet = book.create_sheet("Config")
        config_sheet["A1"] = "Key"
        config_sheet["B1"] = "Value"
        saveFile()
        print("The 'Config' sheet did not exist and has been created with headers.")

    # Iterate over the rows and collect the key-value pair
    for row in config_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        key, value = row
        if key == key_name:
            return {key: value}  # Return key-value pair as a dictionary

    # If key not found
    print(f"Config for '{key_name}' not found in 'Config' sheet.")
    return None


def read_float32(register: int, retries: int = 5) -> float | None:
    """
    Read two consecutive registers and combine them into a 32-bit float.
    Retries the read operation up to a specified number of times if it fails.
    """
    global sameLoop
    for attempt in range(retries):
        try:
            # Read two consecutive registers (32 bits)
            result = client.read_holding_registers(register, 2, unit=1)
            if not result.isError():
                # Combine the two 16-bit registers into a 32-bit float
                high, low = result.registers
                packed = struct.pack(">HH", low, high)
                return struct.unpack(">f", packed)[0]
            else:
                sameLoop = result
        except Exception:
            sameLoop = "Modbus Error"

    return None


def sort_excel_sheets() -> None:
    """
    Read Excel file from directory, sort each sheet by 'Date' and 'S.No',
    write sorted data back while excluding specified sheets.
    """
    # Construct the full file path
    input_file = os.path.join(excel_dir, excel_name)

    # Excluded sheet names
    excluded_sheets = {"Config", "ModelAmp", "Users"}

    try:
        # Get all sheet names
        sheet_data = pd.ExcelFile(input_file)
        sheet_names = sheet_data.sheet_names

        with pd.ExcelWriter(
            input_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            for sheet_name in sheet_names:
                # Skip excluded sheets
                if sheet_name in excluded_sheets:
                    continue

                # Read the current sheet
                sheet_df = pd.read_excel(input_file, sheet_name=sheet_name)

                # Check if required columns exist
                if {"Date", "S.No"}.issubset(sheet_df.columns):
                    # Sort by 'Date' and 'S.No'
                    sorted_df = sheet_df.sort_values(
                        by=["Date", "S.No"], ascending=True, na_position="last"
                    )
                    # Write the sorted data back to the sheet
                    sorted_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    display(
                        f"Skipping sheet '{sheet_name}' as required columns are missing."
                    )
    except FileNotFoundError:
        display(f"File not found: {input_file}. Please ensure the file exists.")
    except Exception as e:
        display(f"An error occurred: {e}")


def check_stop_condition() -> None:
    """Check for ESC key press to stop automation."""
    if keyboard.is_pressed("esc"):
        display("\nStop signal detected. Stopping automation...", "white")
        sort_excel_sheets()
        sys.exit()


def connect_verify() -> None:
    """Verify and establish connection to Modbus server."""
    while True:
        check_stop_condition()
        if not client.connect():
            display("Unable to connect to Modbus server", "yellow", True)
            sleep(3)
        else:
            while True:
                result = read_float32(101 - 1)
                if (
                    result is None
                    and sameLoop is not None
                    and "Modbus Error" in str(sameLoop)
                ):
                    display("Waiting for the connection...", "magenta", True)
                    sleep(5)
                elif "continue" in str(sameLoop):
                    display("Waiting for next machine...", "magenta", True)
                    sleep(5)
                else:
                    break
                # Check for stop condition at each attempt
                check_stop_condition()
            break


def display(txt: str, color: str = "white", same_line: bool = False) -> None:
    """Display colored text in terminal."""
    global colors

    # Get the chosen color or default to Fore.WHITE if the color is not found
    chosen_color = colors.get(color.lower(), Fore.WHITE)
    if same_line:
        print(f"\r{chosen_color + txt}", end="", flush=True)
    else:
        # Print text in specified font and color
        print(chosen_color + txt)


def displayTable(input_data: dict) -> None:
    """
    Display data in a table format with a single separation line between rows.

    Parameters:
        input_data (dict): A dictionary of key-value pairs.
    """
    global partial_row, accumulated_rows

    # Add the new key-value pair(s) to the partial row
    partial_row.update(input_data)

    # Check if the partial row is complete (12 key-value pairs)
    if len(partial_row) >= 12:
        # Add the completed row to the list of accumulated rows
        accumulated_rows.append(partial_row.copy())  # Copy to avoid mutation
        partial_row.clear()  # Reset the partial row

        # Create a PrettyTable object
        table = PrettyTable()
        table.field_names = ["   Parameter   ", "     Value     "]

        test_value = None
        # Add each key-value pair from all rows in accumulated_rows
        for row in accumulated_rows:
            row = {
                key: row.get(key)
                for key in [
                    "Sr.No",
                    "Date",
                    "Time",
                    "MinAmp",
                    "MaxAmp",
                    "Capacitor",
                    "VLN",
                    "Amps",
                    "Watts",
                    "Frequency",
                    "PF",
                    "Reason",
                    "Test",
                ]
            }
            if "Test" in row:
                test_value = row.pop("Test")
            if test_value == "Passed":
                test_value = Fore.GREEN + "Test Passed" + Fore.WHITE
                row.pop("Reason", None)
            elif test_value == "Failed":
                test_value = Fore.RED + "Test Failed" + Fore.WHITE
            for key, value in row.items():
                table.add_row([key, value])

        # Adjust alignment
        table.align["Key"] = "c"
        table.align["Value"] = "c"

        # Get the table as a string with horizontal rules
        table_output = table.get_string(hrules=1)

        # Remove the last line and print the result
        print("\n".join(table_output.splitlines()[:-1]))
        print("+-----------------+-----------------+")
        print(f"|             {test_value}           |")
        print("+-----------------+-----------------+")
        # Reset accumulated rows after printing
        accumulated_rows.clear()


def getAmpModel(serial_name: str) -> list | None:
    """Get amplifier model data for a given serial name."""
    global model_amp_sheet, model_amp_sheet_row
    # Check if the "ModelAmp" sheet exists
    if "ModelAmp" in book.sheetnames:
        # Access the "ModelAmp" sheet
        model_amp_sheet = book["ModelAmp"]
        # Find the last used row in column A (S.No column)
        last_row = 1
        while model_amp_sheet.cell(row=last_row, column=1).value is not None:
            last_row += 1

        # Set excel_row to the next available row
        model_amp_sheet_row = last_row - 1
    else:
        # Create the "ModelAmp" sheet and add headers
        model_amp_sheet = book.create_sheet("ModelAmp")
        model_amp_sheet["A1"] = "Serial"
        model_amp_sheet["B1"] = "Model"
        model_amp_sheet["C1"] = "MaxAmp"
        model_amp_sheet["D1"] = "MinAmp"
        model_amp_sheet["E1"] = "Capacitor"
        saveFile()
        print("The 'ModelAmp' sheet did not exist and has been created with headers.")

    # Iterate over the rows and collect serial, model, maxAmp, and minAmp
    for row in model_amp_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        serial, model, maxAmp, minAmp, capacitor = row
        if serial == serial_name:
            return [model, maxAmp, minAmp, capacitor]

    # If model not found
    display(f"Model {serial_name} not found in 'ModelAmp' sheet.", "cyan", True)
    return None


def checkDuplicate(serial_number: str) -> list:
    """Check if serial number already exists in sheet."""
    global sheet
    # Iterate through the rows in reverse order to get the latest value
    found = False
    qc_value = None
    for row in reversed(
        list(sheet.iter_rows(min_col=1, max_col=11, min_row=2))
    ):
        cell = row[0]  # First column (S.No)
        if cell.value == serial_number:
            found = True
            qc_value = row[10].value
            display(
                f"Found '{serial_number}' in cell {cell.coordinate}     \n",
                "white",
                True,
            )
            break
    return [found, qc_value]


def cellHighlighter(cell_address: str) -> None:
    """Apply green highlight to a cell."""
    green_fill = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    sheet[cell_address].fill = green_fill


def number_to_column(n: int) -> str:
    """Convert column number (1, 2, ..., 27) to Excel column letter (A, B, ..., AA)."""
    col = ""
    while n > 0:
        n -= 1  # Adjust for 1-based indexing
        col = chr(n % 26 + ord("A")) + col
        n //= 26
    return col


def pop_input(title: str, txt: str, input_type: str) -> float | str:
    """Get user input via dialog box with validation."""
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    while True:
        user_inp = simpledialog.askstring(title=title, prompt=txt)

        if input_type == "num":
            try:
                # Ensure the input is a positive number
                if user_inp is not None and float(user_inp) > 0:
                    root.destroy()
                    return float(user_inp)
                else:
                    display("Please enter a positive number.", "white")
            except ValueError:
                display("Invalid input. Please enter a number.", "white")

        elif input_type == "txt":
            # Ensure the input is non-empty
            if user_inp is not None and len(user_inp.strip()) > 0:
                root.destroy()
                return user_inp
            else:
                display("Please enter a non-empty text.", "white")

        # Update root in case of window closure or invalid attempts
        root.update()


def dataTime() -> list[str]:
    """Get current date and time as formatted strings."""
    current = datetime.now()
    # Format the date to dd-mm-yyyy and time to hh:mm:ss
    formatted_date = current.strftime("%d-%m-%Y")
    formatted_time = current.strftime("%H:%M:%S")
    return [formatted_date, formatted_time]


def saveFile() -> None:
    """Save workbook to Excel file."""
    try:
        book.save(os.path.join(excel_dir, excel_name))
    except PermissionError:
        while True:
            try:
                book.save(os.path.join(excel_dir, excel_name))
                break
            except PermissionError:
                display("Please Close The Excel File", "magenta", True)


def heading() -> None:
    """Create sheet headers for the Excel file."""
    headers = {
        "A1": "S.No", "B1": "Model", "C1": "Date", "D1": "Time",
        "E1": "Volt", "F1": "Power Factor", "G1": "Frequency",
        "H1": "Watts", "I1": "Amps", "J1": "Capacitor",
        "K1": "QC", "L1": "Fitter Name"
    }
    for cell, value in headers.items():
        sheet[cell] = value
    try:
        book.save(os.path.join(excel_dir, excel_name))
    except PermissionError:
        display("Please Close The Excel File", "magenta", True)
        heading()


def verify_format(string: str) -> bool:
    """Verify if a string matches the barcode pattern."""
    return bool(re.match(pattern, str(string)))


def is_valid_key(key_name: str) -> bool:
    """Check if the key is alphanumeric or specific barcode symbols."""
    return key_name.isalnum() or key_name in "-_/#"


def decrypt_text(encrypted_text: str, shift: int = 3) -> str:
    """Decrypt text by reversing character shift."""
    decrypted_text = "".join(chr(ord(char) - shift) for char in encrypted_text)
    return decrypted_text


def verify_fitter(input_text: str, workbook, shift: int = 3) -> str | bool:
    """Verify and decrypt fitter name from encrypted input."""
    try:
        # Load the workbook and select the specified sheet
        users = workbook["Users"]
        # Iterate over the rows and find matching fitter
        for row in users.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] == decrypt_text(input_text, shift):
                # Return the decrypted fitter name
                return row[0]
        return False

    except KeyError:
        display("Sheet 'Users' not found in the workbook.")
        return False
    except Exception as e:
        display(f"An error occurred: {e}")
        return False


# Beginning
# Initialize colorama for Windows support
init(autoreset=True)

# Check for updates at the start of the program
threading.Thread(target=check_for_update, daemon=True).start()

color = "red"
colors = {
    "red": Fore.RED,
    "blue": Fore.BLUE,
    "green": Fore.GREEN,
    "yellow": Fore.YELLOW,
    "cyan": Fore.CYAN,
    "magenta": Fore.MAGENTA,
    "white": Fore.WHITE,
}

# Get the chosen color or default to Fore.WHITE if the color is not found
chosen_color = colors.get(color.lower(), Fore.WHITE)
# Get the terminal width and calculate padding
terminal_width = os.get_terminal_size().columns
print("\n" + "=" * terminal_width)
print()
print(f"MotorNoLoadTester By SK96".center(terminal_width))
print(f"Version: {APP_VERSION}".center(terminal_width))
print("\n" + "=" * terminal_width + "\n")

# Path for the configuration file
excel_name = "MotorNoLoadReport.xlsx"
excel_dir = os.path.join(os.environ["USERPROFILE"], "Desktop")

# Variable Initialization
configData = {}
serial_input = ""
FitterName = None
max_amp_value = None
min_amp_value = None
model_amp_sheet = None
model_amp_sheet_row = 1
excel_row = 1
excel_column = 1
sameLoop = "Modbus Error"
config_sheet = None
partial_row = {}
accumulated_rows = []
sheet = None

# Check if the file exists and determine the last row
if not os.path.exists(os.path.join(excel_dir, excel_name)):
    # File does not exist, create a new one
    root = tk.Tk()
    root.withdraw()
    file_create = messagebox.askyesno(
        title="CreateFile",
        message="Excel File does not exist. Would you like to create it?",
    )
    if file_create:
        print(f"Excel File Save Location: {excel_dir}\n")
        book = Workbook()
        # Remove the default sheet if it exists
        if "Sheet" in book.sheetnames:
            default_sheet = book["Sheet"]
            book.remove(default_sheet)
        # Create the "ModelAmp" sheet and add headers
        model_amp_sheet = book.create_sheet("ModelAmp")
        model_amp_sheet["A1"] = "Serial"
        model_amp_sheet["B1"] = "Model"
        model_amp_sheet["C1"] = "MaxAmp"
        model_amp_sheet["D1"] = "MinAmp"
        model_amp_sheet["E1"] = "Capacitor"
        saveFile()

    else:
        print("Program Cancelled")
        sys.exit()
    root.destroy()

# If file exists, load the workbook
book = load_workbook(os.path.join(excel_dir, excel_name))
# Configuration keys
config_keys = [
    "method",
    "port",
    "baudrate",
    "parity",
    "stopbits",
    "bytesize",
    "timeout",
]
# Iterate over configuration keys
for i, key in enumerate(config_keys, start=2):  # Start at row 2 for data
    result = getConfig(key)
    if result is None:
        if key in ("method", "port", "parity"):
            # Prompt user for the missing value
            value = pop_input(key, f"Enter the {key} value", "txt")
        else:
            value = int(pop_input(key, f"Enter the {key} value", "num"))
        config_sheet[f"A{i}"] = key
        config_sheet[f"B{i}"] = value
        saveFile()
        configData[key] = value
    else:
        configData.update(result)  # Add existing key-value pair to the dictionary


# Initialize Modbus client (RTU over serial)
client = ModbusSerialClient(**configData)

# Register addresses for each parameter (e.g., Frequency, VR, etc.)
registers = {"VLN": 141, "PF": 117, "Frequency": 157, "Watts": 101, "Amps": 149}

# Define the regex pattern for barcode format
pattern = r"^[A-Za-z]+\d+[A-Za-z]?-[A-Za-z]\d{2}/\d{3,4}$"

try:
    # Loop to continue asking for input
    while True:
        connect_verify()
        display("Waiting for QR to be read.........\n", "blue", True)
        serial_input = ""  # Reset the old value
        while True:
            check_stop_condition()
            event = keyboard.read_event()
            if event.event_type == keyboard.KEY_DOWN:
                # Ignore "Enter" but treat it as the end of the scan
                if event.name == "enter":
                    break
                if is_valid_key(event.name):
                    serial_input += event.name
        if FitterName is None:
            print(serial_input)
            FitterName = verify_fitter(serial_input, book)
            if FitterName:
                display(f"Fitter Name: {FitterName}")
            else:
                FitterName = None
                serial_input = ""  # Reset the old value
                display("User Not Found, Please contact Admin.")
            print()
            continue
        serial_input = serial_input.upper()
        if not verify_format(serial_input):
            display(f"Pattern Not Matching: {serial_input}")
            print()
            continue
        connect_verify()
        serial_name = (
            serial_input.split("-")[0] if "-" in serial_input else serial_input
        )
        result = getAmpModel(serial_name)
        getSheetLastRow(serial_name)
        duplicate_check = checkDuplicate(serial_input)
        if duplicate_check[0] and duplicate_check[1] == "OK":
            print()
            continue
        if result is not None:
            model_value, max_amp_value, min_amp_value, capacitor_value = result
            display(
                f"The Amp value for model '{serial_input}' are MAX {max_amp_value} MIN {min_amp_value}.\n",
                "white",
                True,
            )
        else:
            root = tk.Tk()
            root.withdraw()
            file_create = messagebox.askyesno(
                title="AddEntry",
                message=f"Model '{serial_name}' does not have an Amp value. Would you like to add it?",
            )
            if file_create:
                model_value = int(
                    pop_input(
                        "ModelValue",
                        f"Enter the Model Number for '{serial_name}'",
                        "num",
                    )
                )
                max_amp_value = float(
                    pop_input(
                        "MaxAmpValue",
                        f"Enter the Max Amp Value for '{serial_name}'",
                        "num",
                    )
                )
                min_amp_value = float(
                    pop_input(
                        "MinAmpValue",
                        f"Enter the Min Amp Value for '{serial_name}'",
                        "num",
                    )
                )
                capacitor_value = int(
                    pop_input(
                        "CapacitorValue",
                        f"Enter the Capacitor Value for '{serial_name}'",
                        "num",
                    )
                )
                model_amp_sheet[f"A{model_amp_sheet_row + 1}"] = serial_name
                model_amp_sheet[f"B{model_amp_sheet_row + 1}"] = model_value
                model_amp_sheet[f"C{model_amp_sheet_row + 1}"] = max_amp_value
                model_amp_sheet[f"D{model_amp_sheet_row + 1}"] = min_amp_value
                model_amp_sheet[f"E{model_amp_sheet_row + 1}"] = capacitor_value
                saveFile()
            else:
                display(
                    f"Model {serial_name} not found:- Test Cancelled: {serial_input}\n",
                    "yellow",
                    True,
                )
                continue
        sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = serial_input
        excel_column += 1
        sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = model_value
        excel_column += 1
        date_time = dataTime()
        sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = date_time[0]
        excel_column += 1
        sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = date_time[1]
        excel_column += 1
        displayTable(
            {
                "Sr.No": serial_input,
                "Date": date_time[0],
                "Time": date_time[1],
                "MinAmp": min_amp_value,
                "MaxAmp": max_amp_value,
                "Capacitor": capacitor_value,
            }
        )
        amp_value = None
        for param, register in registers.items():
            value = read_float32(register - 1)
            if value is not None:
                sameLoop = "continue"
                value = f"{value:.2f}"
                sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = value
                excel_column += 1
                displayTable({param: value})
                if param == "Amps":
                    sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = capacitor_value
                    excel_column += 1
                    amp_value = param
                    if float(value) >= min_amp_value and float(value) <= max_amp_value:
                        displayTable({"Test": "Passed"})
                        sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = "OK"
                        if duplicate_check[0]:
                            cellHighlighter(
                                f"{number_to_column(excel_column)}{excel_row + 1}"
                            )
                        excel_column += 1
                    else:
                        if float(value) <= min_amp_value:
                            displayTable(
                                {
                                    "Reason": Fore.RED + "LowAmps" + Fore.WHITE,
                                    "Test": "Failed",
                                }
                            )
                        elif float(value) >= max_amp_value:
                            displayTable(
                                {
                                    "Reason": Fore.RED + "HighAmps" + Fore.WHITE,
                                    "Test": "Failed",
                                }
                            )
                        else:
                            displayTable(
                                {
                                    "Reason": Fore.RED + "Amps" + Fore.WHITE,
                                    "Test": "Failed",
                                }
                            )
                        sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = "NOT OK"
                        excel_column += 1
            else:
                display(f"Could not read value for {param}", "red")
        print()
        if FitterName is not None:
            sheet[f"{number_to_column(excel_column)}{excel_row + 1}"] = FitterName
            excel_column += 1
        if value is None or float(value) <= 0:
            excel_column = 1
            continue
        excel_row += 1
        excel_column = 1
        saveFile()
        sleep(5)
except Exception as e:
    display(Style.BRIGHT + f"Error on the Program Contact the Admin: {e}", "red")

# Disconnect client
client.close()
