from pymodbus.client.sync import ModbusSerialClient
import struct
from openpyxl import load_workbook, Workbook
from time import sleep
import tkinter as tk
from tkinter import simpledialog
import keyboard
import os
from datetime import datetime
from tkinter import messagebox
from colorama import Fore, Style, init
import re, random
from openpyxl.styles import PatternFill
import sys
from prettytable import PrettyTable # type: ignore


# watchmedo auto-restart --patterns="*.py" --recursive -- python d:/Dev/MFM/MFM367_Data.py
def getConfig(Key):
    global config_sheet
    # Check if the "Config" sheet exists
    if "Config" in book.sheetnames:
        # Access the "Config" sheet
        config_sheet = book["Config"]
    else:
        # Create the "Config" sheet and add headers
        config_sheet = book.create_sheet("Config")
        config_sheet["A1"] = "Key"
        config_sheet["B1"] = "Value"
        saveFile()
        print("The 'Config' sheet did not exist and has been created with headers.")

    # Iterate over the rows and collect the key-value pair
    for row in config_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        key, value = row
        if key == Key:
            return {key: value}  # Return key-value pair as a dictionary

    # If key not found
    print(f"Config for '{Key}' not found in 'Config' sheet.")
    return None


def read_float32(register, retries=5):
    """
    Read two consecutive registers and combine them into a 32-bit float.
    Retries the read operation up to a specified number of times if it fails.
    """
    global sameLoop
    for attempt in range(retries):
        try:
            return 4.4567869 # removee
            # Read two consecutive registers (32 bits)
            result = client.read_holding_registers(register, 2, unit=1)
            if not result.isError():
                # Combine the two 16-bit registers into a 32-bit float
                high, low = result.registers
                packed = struct.pack('>HH', low, high)
                return struct.unpack('>f', packed)[0]
            else:
                # display(f"Error reading register {register}: {result}", "red")  # debug
                sameLoop = result
        except Exception as e:
            # display(Style.BRIGHT + f"Exception reading register {register}: {e}", "red")  # debug
            sameLoop = "Modbus Error"
        # display(f"Retrying... ({attempt + 1}/{retries})", "cyan")  # debug
    
    # display(f"Failed to read register {register} after {retries} attempts.", "red", True)  # debug
    return None


# Function to check and handle the stop condition
def check_stop_condition():
    # Check for 'ctrl+shift+alt' key combination in a single string
    if keyboard.is_pressed('esc'):
        display("\nStop signal detected. Stopping automation...", "white")
        sys.exit()


def connect_verify():
    while True:
        check_stop_condition()
        if not client.connect():
            display("Unable to connect to Modbus server", "yellow", True)
            sleep(3)
            break # removee
        else:
            while True:
                result = read_float32(101-1)
                if result is None and sameLoop is not None and "Modbus Error" in str(sameLoop):
                    display("Waiting for the connection...", "magenta", True)
                    sleep(5)
                    break # removee
                elif "continue" in str(sameLoop):
                    display("Waiting for next machine...", "magenta", True)
                    sleep(5)
                    break # removee
                else:
                    break
                # Check for stop condition at each attempt
                check_stop_condition()
            break


def display(txt, color = "white", sameLine = False):
    global colors
    
    # Get the chosen color or default to Fore.WHITE if the color is not found
    chosen_color = colors.get(color.lower(), Fore.WHITE)
    if sameLine:
        print(f"\r{chosen_color + txt}", end="", flush=True)
    else:
        # Print text in specified font and color
        print(chosen_color + txt)


def displayTable(input_data):
    """
    Displays data in a table format with a single separation line between rows.

    Parameters:
        input_data (dict): A dictionary of key-value pairs.
    """
    global partial_row, accumulated_rows

    # Add the new key-value pair(s) to the partial row
    partial_row.update(input_data)

    # Check if the partial row is complete (12 key-value pairs)
    if len(partial_row) >= 12:
        # Add the completed row to the list of accumulated rows
        accumulated_rows.append(partial_row.copy())  # Copy to avoid mutation
        partial_row.clear()  # Reset the partial row

        # Create a PrettyTable object
        table = PrettyTable()
        table.field_names = ["   Parameter   ", "     Value     "]

        test_value = None
        # Add each key-value pair from all rows in accumulated_rows
        for row in accumulated_rows:
            row = {key: row.get(key) for key in ["Sr.No", "Date", "Time", "MinAmp", "MaxAmp", "Capacitor", "VLN", "Amps", "Watts", "Frequency", "PF", "Reason", "Test"]}
            if "Test" in row:
                test_value = row.pop("Test")
            if test_value == "Passed":
                test_value = Fore.GREEN + "Test Passed" + Fore.WHITE
                row.pop("Reason")
            elif test_value == "Failed":
                test_value = Fore.RED + "Test Failed" + Fore.WHITE
            for key, value in row.items():
                table.add_row([key, value])

        # Adjust alignment
        table.align["Key"] = "c"
        table.align["Value"] = "c"

        # Get the table as a string with horizontal rules
        table_output = table.get_string(hrules=1)

        # Remove the last line and print the result
        print("\n".join(table_output.splitlines()[:-1]))
        print("+-----------------+-----------------+")
        print(f"|             {test_value}           |")
        print("+-----------------+-----------------+")
        # Reset accumulated rows after printing
        accumulated_rows.clear()


def getAmpModel(serial_name):
    global model_amp_sheet, model_amp_sheet_row
    # Check if the "ModelAmp" sheet exists
    if "ModelAmp" in book.sheetnames:
        # Access the "ModelAmp" sheet
        model_amp_sheet = book["ModelAmp"]
        # Find the last used row in column A (S.No column)
        last_row = 1
        while model_amp_sheet.cell(row=last_row, column=1).value is not None:
            last_row += 1

        # Set excel_row to the next available row
        model_amp_sheet_row = last_row - 1
    else:
        # Create the "ModelAmp" sheet and add headers
        model_amp_sheet = book.create_sheet("ModelAmp")
        model_amp_sheet["A1"] = "Serial"
        model_amp_sheet["B1"] = "Model"
        model_amp_sheet["C1"] = "MaxAmp"
        model_amp_sheet["D1"] = "MinAmp"
        model_amp_sheet["E1"] = "Capacitor"
        saveFile()
        print("The 'ModelAmp' sheet did not exist and has been created with headers.")

    # Iterate over the rows and collect serial, model, maxAmp, and minAmp
    for row in model_amp_sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        serial, model, maxAmp, minAmp, capacitor = row
        if serial == serial_name:
            return [model, maxAmp, minAmp, capacitor]

    # If model not found
    display(f"Model {serial_name} not found in 'ModelAmp' sheet.", "cyan", True)
    return None


def checkDuplicate(serial_number):
    global sheet, excel_row
    # Iterate through the rows in reverse order to get the latest value
    found = False
    QCValue = None
    for row in reversed(list(sheet.iter_rows(min_col=1, max_col=11, min_row=2))):  # Reverse order
        cell = row[0]  # First column (S.No)
        if cell.value == serial_number:
            found = True
            QCValue = row[10].value
            display(f"Found '{serial_number}' in cell {cell.coordinate}     \n", "white", True)
            break
    return [found, QCValue]


def cellHighlighter(cell_address):
    # Apply green highlight to the cell
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    sheet[cell_address].fill = green_fill


def number_to_column(n):
    """Convert a column number (e.g., 1, 2, ..., 27) to an Excel column letter (e.g., A, B, ..., AA)."""
    col = ''
    while n > 0:
        n -= 1  # Adjust for 1-based indexing
        col = chr(n % 26 + ord('A')) + col
        n //= 26
    return col


def pop_input(title, txt, input_type):
    global excel_row, excel_column
    ROOT = tk.Tk()
    ROOT.withdraw()  # Hide the root window
    while True:
        user_inp = simpledialog.askstring(title=title, prompt=txt)
        
        if input_type == "num":
            try:
                # Ensure the input is a positive integer
                if user_inp is not None and float(user_inp) > 0:
                    ROOT.destroy()
                    return float(user_inp)  # Returning as integer
                else:
                    display("Please enter a positive number.", "white")
            except ValueError:
                display("Invalid input. Please enter a number.", "white")
                
        elif input_type == "txt":
            # Ensure the input is non-empty
            if user_inp is not None and len(user_inp.strip()) > 0:
                ROOT.destroy()
                return user_inp
            else:
                display("Please enter a non-empty text.", "white")
                
        # Destroy ROOT in case of window closure or invalid attempts
        ROOT.update()


def dataTime():
    global excel_row, excel_column
    # Get the current date
    current = datetime.now()
    # Format the date to dd-mm-yyyy
    formatted_date = current.strftime('%d-%m-%Y')
    # Format the time to hh:mm:ss
    formatted_time = current.strftime('%H:%M:%S')
    return [formatted_date, formatted_time]

def saveFile():
    try:
        book.save(excel_dir + "\\" + excel_name)
    except PermissionError:
        while True:
            try:
                book.save(excel_dir + "\\" + excel_name)
                break
            except PermissionError:
                display("Please Close The Excel File", "magenta", True)

def heading():
    sheet['A1'] = 'S.No'
    sheet['B1'] = 'Model'
    sheet['C1'] = 'Date'
    sheet['D1'] = 'Time'
    sheet['E1'] = 'Volt'
    sheet['F1'] = 'Power Factor'
    sheet['G1'] = 'Frequency'
    sheet['H1'] = 'Watts'
    sheet['I1'] = 'Amps'
    sheet['J1'] = 'Capacitor'
    sheet['K1'] = 'QC'
    sheet['L1'] = 'Fitter Name'
    try:
        book.save(excel_dir + "\\" + excel_name)
    except PermissionError:
        display("Please Close The Excel File", "magenta", True)
        heading()


# Function to verify if a string matches the pattern
def verify_format(string):
    return bool(re.match(pattern, str(string)))


def is_valid_key(key_name):
    """Check if the key is alphanumeric or specific barcode symbols."""
    return key_name.isalnum() or key_name in "-_/"


# Beginning
# Initialize colorama for Windows support
init(autoreset=True)

color = "red"
colors = {
    "red": Fore.RED,
    "blue": Fore.BLUE,
    "green": Fore.GREEN,
    "yellow": Fore.YELLOW,
    "cyan": Fore.CYAN,
    "magenta": Fore.MAGENTA,
    "white": Fore.WHITE
}

# Get the chosen color or default to Fore.WHITE if the color is not found
chosen_color = colors.get(color.lower(), Fore.WHITE)
# Get the terminal width and calculate padding
terminal_width = os.get_terminal_size().columns
print("\n" + "=" * terminal_width)
print()
print(f"MotorNoLoadTester By SK96".center(terminal_width))
print("\n" + "=" * terminal_width + "\n")

# Path for the configuration file
excel_name = 'MotorNoLoadReport.xlsx'
excel_dir = os.path.join(os.environ['USERPROFILE'], 'Desktop')
# excel_dir = "D:\\Dev\\MFM"

# Variable Initialization
configData = {}
SrNoUSER_INP = ""
FitterName = None
FitterName = "SM" # removee
MaxAmpValue = None
MinAmpValue = None
model_amp_sheet = None
model_amp_sheet_row = 1
excel_row = 1
excel_column = 1
sameLoop = "Modbus Error"
config_sheet = None
partial_row = {}
accumulated_rows = []

# Check if the file exists and determine the last row
if os.path.exists(os.path.join(excel_dir, excel_name)):
    # If file exists, load the workbook
    book = load_workbook(os.path.join(excel_dir, excel_name))

    # Explicitly target the "Data" sheet
    if "Data" in book.sheetnames:
        sheet = book["Data"]
    else:
        # Handle the case where "Data" sheet does not exist
        sheet = book.create_sheet("Data")
        heading()

    # Find the last used row in column A (S.No column)
    last_row = 1
    while sheet.cell(row=last_row, column=1).value is not None:
        last_row += 1

    # Set excel_row to the next available row
    excel_row = last_row - 1
else:
    # File does not exist, create a new one
    root = tk.Tk()
    root.withdraw()
    fileCreate = messagebox.askyesno(title="CreateFile", message="Excel File does not exist. Would you like to create it?")
    if fileCreate:
        print('Excel File Save Location: ' + excel_dir + '\n')
        book = Workbook()

        # Remove the default sheet if it exists
        if "Sheet" in book.sheetnames:
            default_sheet = book["Sheet"]
            book.remove(default_sheet)

        # Create the "Data" sheet explicitly
        if "Data" in book.sheetnames:
            sheet = book["Data"]
        else:
            sheet = book.create_sheet("Data")

        # Add headings or initialization if needed
        heading()

    else:
        print("Program Cancelled")
        sys.exit()
    root.destroy()

# Configuration keys
config_keys = ['method', 'port', 'baudrate', 'parity', 'stopbits', 'bytesize', 'timeout']
# Iterate over configuration keys
for i, key in enumerate(config_keys, start=2):  # Start at row 2 for data
    result = getConfig(key)
    if result is None:
        if key in ('method', 'port', 'parity'):
            # Prompt user for the missing value
            value = pop_input(key, f"Enter the {key} value", "txt")
        else:
            value = int(pop_input(key, f"Enter the {key} value", "num"))
        config_sheet[f"A{i}"] = key
        config_sheet[f"B{i}"] = value
        saveFile()
        configData[key] = value
    else:
        configData.update(result)  # Add existing key-value pair to the dictionary


# Initialize Modbus client (RTU over serial)
client = ModbusSerialClient(**configData)
# client = ModbusSerialClient(
#     method='rtu',
#     port='COM3',
#     baudrate=9600,
#     parity='E',
#     stopbits=1,
#     bytesize=8,
#     timeout=1)

# Register addresses for each parameter (e.g., Frequency, VR, etc.)
registers = {
    "VLN": 141,
    "PF": 117,
    "Frequency": 157,
    "Watts": 101,
    "Amps": 149
}

# Define the regex pattern according to the updated requirements
pattern = r"^[A-Za-z]\d+(\.\d{1})?[A-Za-z]-\d{2}/\d{4}$"

# Test strings
test_strings = ["F4R-24/0001", "F1.5R-24/0001", "X4.2Y-25/0001", "A5Z-245/0001", "B1.5A-12/0001", "C10R-99/0001"] # removee

try:
    # Loop to continue asking for input
    while True:
        connect_verify()
        display("Waiting for QR to be read.........\n", "blue", True)
        SrNoUSER_INP = ""  # Reset the old Value
        while True:
            check_stop_condition()
            SrNoUSER_INP = random.choice(test_strings) # removee
            break # removee
            event = keyboard.read_event()
            if event.event_type == keyboard.KEY_DOWN:
                # Ignore "Enter" but treat it as the end of the scan
                if event.name == "enter":
                    break
                if is_valid_key(event.name):
                    SrNoUSER_INP += str(event.name.upper()) if event.name.islower() else event.name

        if not verify_format(SrNoUSER_INP):
            display(f"Pattern Not Matching: {SrNoUSER_INP}")
            print()
            continue
        connect_verify()
        duplicate_check = checkDuplicate(SrNoUSER_INP)
        if duplicate_check[0] and duplicate_check[1] == "OK":
            print()
            continue
        serial_name = SrNoUSER_INP.split('-')[0] if '-' in SrNoUSER_INP else SrNoUSER_INP
        result = getAmpModel(serial_name)
        if result is not None:
            ModelValue, MaxAmpValue, MinAmpValue, CapacitorValue = result
            display(f"The Amp value for model '{SrNoUSER_INP}' are MAX {MaxAmpValue} MIN {MinAmpValue}.\n", "white", True)
        else:
            root = tk.Tk()
            root.withdraw()
            fileCreate = messagebox.askyesno(title="AddEntry", message=f"Model '{serial_name}' does not have an Amp value. Would you like to add it?")
            if fileCreate:
                ModelValue = int(pop_input("ModelValue", f"Enter the Model Number for '{serial_name}'", "num"))
                MaxAmpValue = float(pop_input("MaxAmpValue", f"Enter the Max Amp Value for '{serial_name}'", "num"))
                MinAmpValue = float(pop_input("MinAmpValue", f"Enter the Min Amp Value for '{serial_name}'", "num"))
                CapacitorValue = int(pop_input("CapacitorValue", f"Enter the Capacitor Value for '{serial_name}'", "num"))
                model_amp_sheet['A'+str(model_amp_sheet_row + 1)] = serial_name
                model_amp_sheet['B'+str(model_amp_sheet_row + 1)] = ModelValue
                model_amp_sheet['C'+str(model_amp_sheet_row + 1)] = MaxAmpValue
                model_amp_sheet['D'+str(model_amp_sheet_row + 1)] = MinAmpValue
                model_amp_sheet['E'+str(model_amp_sheet_row + 1)] = CapacitorValue
                saveFile()
            else:
                display(f"Model {serial_name} not found:- Test Cancelled: {SrNoUSER_INP}\n", "yellow", True)
                continue
        sheet[number_to_column(excel_column) + str(excel_row + 1)] = SrNoUSER_INP
        excel_column += 1
        sheet[number_to_column(excel_column) + str(excel_row + 1)] = ModelValue
        excel_column += 1
        if FitterName == None:
            FitterName =  pop_input("Enter Your Name", "Fitter Name: ", "txt")
        dateTime = dataTime()
        # display(str(dateTime) + "                 \n", "white", True)
        sheet[number_to_column(excel_column) + str(excel_row + 1)] = dateTime[0]
        excel_column += 1
        sheet[number_to_column(excel_column) + str(excel_row + 1)] = dateTime[1]
        excel_column += 1
        displayTable({"Sr.No": SrNoUSER_INP, "Date": dateTime[0], "Time": dateTime[1], "MinAmp": MinAmpValue, "MaxAmp": MaxAmpValue, "Capacitor": CapacitorValue})
        AmpValue = None
        for param, register in registers.items():
            value = read_float32(register - 1)
            if value is not None:
                sameLoop = "continue"
                value = "{:.2f}".format(value)
                sheet[number_to_column(excel_column) + str(excel_row + 1)] = value
                excel_column += 1
                displayTable({param: value})
                # print(f"{param}: {value}")
                if param == "Amps":
                    sheet[number_to_column(excel_column) + str(excel_row + 1)] = CapacitorValue
                    excel_column += 1
                    AmpValue = param
                    if float(value) >= MinAmpValue and float(value) <= MaxAmpValue:
                        displayTable({"Test": "Passed" })
                        # display(f"Test Passed: {SrNoUSER_INP}", "green")
                        sheet[number_to_column(excel_column) + str(excel_row + 1)] = "OK"
                        if duplicate_check[0]:
                            cellHighlighter(number_to_column(excel_column) + str(excel_row + 1))
                        excel_column += 1
                    else:
                        if float(value) <= MinAmpValue:
                            displayTable({"Reason": Fore.RED + "LowAmps" + Fore.WHITE, "Test": "Failed" })
                        elif float(value) >= MaxAmpValue:
                            displayTable({"Reason": Fore.RED + "HighAmps" + Fore.WHITE, "Test": "Failed" })
                        else:
                            displayTable({"Reason": Fore.RED + "Amps" + Fore.WHITE, "Test": "Failed" })
                        # displayTable({"QC": "Failed"})
                        sheet[number_to_column(excel_column) + str(excel_row + 1)] = "NOT OK"
                        excel_column += 1
            else:
                display(f"Could not read value for {param}", "red")
        print()
        if FitterName is not None:
            sheet[number_to_column(excel_column) + str(excel_row + 1)] = FitterName
            excel_column += 1
        if value is None or float(value) <= 0:
            excel_column = 1
            continue
        excel_row += 1
        excel_column = 1
        saveFile()
        sleep(5)
except Exception as e:
    display(Style.BRIGHT + f"Error on the Program Contact the Admin: {e}", "red")

# Disconnect client
client.close()